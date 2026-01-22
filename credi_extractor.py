#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CREDI Extractor - Application d'extraction de commandes de repas
Version Python avec interface graphique
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import sqlite3
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import threading

# Imports pour la lecture de fichiers
try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None

try:
    import PyPDF2
    import pdfplumber
except ImportError:
    PyPDF2 = None
    pdfplumber = None


class SchoolNameNormalizer:
    """Gestionnaire de normalisation des noms d'écoles"""
    
    def __init__(self):
        self.mappings: Dict[str, str] = {}
        self.replacements: List[Tuple[str, str]] = [
            (r'\bSTE\b', 'SAINTE'),
            (r'\bST\b', 'SAINT'),
            (r'MATERNELLEELEMENTAIRE', 'MATERNELLE ELEMENTAIRE'),
            (r'ELEMENTAIREMATERNELLE', 'ELEMENTAIRE MATERNELLE'),
            (r'\bMAT\b', 'MATERNELLE'),
            (r'\bELEM\b', 'ELEMENTAIRE'),
            (r'\bELE\b', 'ELEMENTAIRE'),
            (r'\bECOLE\b', ''),
            (r'\bCANTINE SCOLAIRE\b', ''),
            (r'\bRESTAURATION\b', ''),
            (r'\+', ' '),
            (r'\bET\b', ''),
            (r'\s+', ' '),
        ]
        self.found_schools = set()
        self.normalization_stats = {}
    
    def load_mapping_file(self, filepath: str) -> bool:
        """Charge un fichier de mapping JSON"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Charger les replacements si disponibles
            if 'replacements' in data:
                self.replacements = [(r['pattern'], r['replace']) for r in data['replacements']]
            
            # Charger les mappings
            if 'mappings' in data:
                self.mappings = data['mappings']
            
            return True
        except Exception as e:
            print(f"Erreur lors du chargement du mapping: {e}")
            return False
    
    def normalize(self, school_name: str) -> str:
        """Normalise un nom d'école"""
        if not school_name:
            return "ECOLE_INCONNUE"
        
        original = school_name
        
        # Normalisation de base
        normalized = school_name.strip().upper()
        
        # Supprimer les accents
        normalized = self._remove_accents(normalized)
        
        # Vérifier si un mapping direct existe
        if normalized in self.mappings:
            result = self.mappings[normalized]
            self.found_schools.add(original)
            self._record_normalization(original, result)
            return result
        
        # Appliquer les replacements
        for pattern, replacement in self.replacements:
            normalized = re.sub(pattern, replacement, normalized)
        
        # Nettoyer les espaces multiples
        normalized = ' '.join(normalized.split())
        normalized = normalized.strip()
        
        if not normalized:
            normalized = "ECOLE_INCONNUE"
        
        self.found_schools.add(original)
        self._record_normalization(original, normalized)
        
        return normalized
    
    def _remove_accents(self, text: str) -> str:
        """Supprime les accents d'un texte"""
        import unicodedata
        return ''.join(
            c for c in unicodedata.normalize('NFD', text)
            if unicodedata.category(c) != 'Mn'
        )
    
    def _record_normalization(self, original: str, normalized: str):
        """Enregistre une normalisation pour les statistiques"""
        if original != normalized:
            if normalized not in self.normalization_stats:
                self.normalization_stats[normalized] = []
            if original not in self.normalization_stats[normalized]:
                self.normalization_stats[normalized].append(original)
    
    def get_normalization_report(self) -> str:
        """Génère un rapport de normalisation"""
        if not self.normalization_stats:
            return "Aucune normalisation effectuée"
        
        report = f"✅ {len(self.normalization_stats)} noms d'écoles consolidés\n\n"
        
        for canonical, variations in sorted(self.normalization_stats.items()):
            if len(variations) > 1:
                report += f"📍 {canonical}:\n"
                for var in variations:
                    report += f"   ← {var}\n"
                report += "\n"
        
        return report
    
    def export_found_schools(self, output_dir: str):
        """Exporte la liste des écoles trouvées"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Export JSON
        json_file = os.path.join(output_dir, f"ecoles_trouvees_{timestamp}.json")
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump({
                "date": datetime.now().isoformat(),
                "total_schools": len(self.found_schools),
                "schools": sorted(list(self.found_schools))
            }, f, ensure_ascii=False, indent=2)
        
        # Export TXT
        txt_file = os.path.join(output_dir, f"ecoles_trouvees_{timestamp}.txt")
        with open(txt_file, 'w', encoding='utf-8') as f:
            f.write(f"Liste des écoles trouvées - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total: {len(self.found_schools)} écoles\n")
            f.write("=" * 80 + "\n\n")
            for school in sorted(self.found_schools):
                f.write(f"{school}\n")
        
        return json_file, txt_file


class MealOrderExtractor:
    """Extracteur de données de commandes de repas"""
    
    def __init__(self, normalizer: SchoolNameNormalizer):
        self.normalizer = normalizer
        self.db_connection = None
        self.extracted_data = []
    
    def create_database(self, db_path: str):
        """Crée la base de données SQLite"""
        self.db_connection = sqlite3.connect(db_path)
        cursor = self.db_connection.cursor()
        
        # Table principale des commandes
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS commandes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ecole TEXT NOT NULL,
                ecole_normalized TEXT NOT NULL,
                agent TEXT,
                date_commande TEXT,
                jour TEXT,
                jour_semaine TEXT,
                type_repas TEXT,
                categorie TEXT,
                quantite INTEGER,
                source_fichier TEXT,
                date_extraction TEXT
            )
        ''')
        
        # Table de métadonnées
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS metadata (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')
        
        cursor.execute('''
            INSERT OR REPLACE INTO metadata (key, value) VALUES (?, ?)
        ''', ('created_at', datetime.now().isoformat()))
        
        cursor.execute('''
            INSERT OR REPLACE INTO metadata (key, value) VALUES (?, ?)
        ''', ('version', 'credi_extractor_python_v1'))
        
        self.db_connection.commit()
    
    def extract_from_excel(self, file_path: str) -> List[Dict]:
        """Extrait les données d'un fichier Excel"""
        if not openpyxl:
            raise ImportError("openpyxl n'est pas installé. Installez-le avec: pip install openpyxl")
        
        results = []
        wb = load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Détecter l'école et l'agent
            school_name = None
            agent_name = None
            period = None
            
            for row in sheet.iter_rows(max_row=20, values_only=True):
                for cell_value in row:
                    if cell_value and isinstance(cell_value, str):
                        # Rechercher le nom de l'école
                        if 'RESTAURANT' in cell_value.upper() or 'ECOLE' in cell_value.upper():
                            match = re.search(r':\s*(.+)', cell_value)
                            if match:
                                school_name = match.group(1).strip()
                        
                        # Rechercher l'agent
                        if 'AGENT' in cell_value.upper():
                            match = re.search(r':\s*(.+)', cell_value)
                            if match:
                                agent_name = match.group(1).strip()
                        
                        # Rechercher la période
                        if 'PERIODE' in cell_value.upper() or 'SEMAINE' in cell_value.upper():
                            period = cell_value
            
            # Si pas d'école trouvée, utiliser le nom du fichier
            if not school_name:
                school_name = Path(file_path).stem
            
            normalized_school = self.normalizer.normalize(school_name)
            
            # Extraire les données du tableau
            data_rows = self._extract_table_data(sheet)
            
            for row_data in data_rows:
                row_data.update({
                    'ecole': school_name,
                    'ecole_normalized': normalized_school,
                    'agent': agent_name,
                    'periode': period,
                    'source_fichier': os.path.basename(file_path)
                })
                results.append(row_data)
        
        return results
    
    def extract_from_pdf(self, file_path: str) -> List[Dict]:
        """Extrait les données d'un fichier PDF"""
        if not pdfplumber:
            raise ImportError("pdfplumber n'est pas installé. Installez-le avec: pip install pdfplumber")
        
        results = []
        
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                
                if not text:
                    continue
                
                # Extraire les informations de l'en-tête
                school_name = None
                agent_name = None
                
                lines = text.split('\n')
                for line in lines:
                    if 'RESTAURANT' in line.upper() or 'ECOLE' in line.upper():
                        match = re.search(r':\s*(.+)', line)
                        if match:
                            school_name = match.group(1).strip()
                    
                    if 'AGENT' in line.upper():
                        match = re.search(r':\s*(.+)', line)
                        if match:
                            agent_name = match.group(1).strip()
                
                if not school_name:
                    school_name = Path(file_path).stem
                
                normalized_school = self.normalizer.normalize(school_name)
                
                # Extraire les données numériques
                # Pattern pour détecter les catégories et quantités
                patterns = {
                    'Enfants': r'Enfants.*?(\d+)\s+(\d+)\s+(\d+)\s+(\d+)',
                    'Standard': r'Standard\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)',
                    'Sans porc': r'Sans porc\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)',
                    'Végétarien': r'Végétarien\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)',
                    'Halal': r'Halal\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)',
                }
                
                jours = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']
                
                for categorie, pattern in patterns.items():
                    match = re.search(pattern, text)
                    if match:
                        quantites = [int(q) for q in match.groups()]
                        for i, jour in enumerate(jours[:len(quantites)]):
                            if quantites[i] > 0:
                                results.append({
                                    'ecole': school_name,
                                    'ecole_normalized': normalized_school,
                                    'agent': agent_name,
                                    'jour_semaine': jour,
                                    'categorie': categorie,
                                    'quantite': quantites[i],
                                    'source_fichier': os.path.basename(file_path)
                                })
        
        return results
    
    def _extract_table_data(self, sheet) -> List[Dict]:
        """Extrait les données du tableau Excel"""
        data = []
        
        # Rechercher la ligne d'en-tête avec les jours
        header_row = None
        jours_colonnes = {}
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_idx, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    cell_upper = cell.upper().strip()
                    if cell_upper in ['LUNDI', 'MARDI', 'MERCREDI', 'JEUDI', 'VENDREDI']:
                        if header_row is None:
                            header_row = row_idx
                        jours_colonnes[col_idx] = cell_upper
            
            if header_row and row_idx > header_row:
                break
        
        if not jours_colonnes:
            return data
        
        # Extraire les données après l'en-tête
        categories = ['Standard', 'Sans porc', 'Végétarien', 'Halal']
        current_type = None
        
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            first_cell = str(row[0] or '').strip()
            
            # Détecter le type (Enfants/Adultes/Maternelle)
            if first_cell.upper() in ['ENFANTS', 'ADULTES', 'MATERNELLE']:
                current_type = first_cell.title()
                continue
            
            # Détecter la catégorie
            if first_cell in categories:
                categorie = first_cell
                
                # Extraire les quantités pour chaque jour
                for col_idx, jour in jours_colonnes.items():
                    try:
                        quantite = row[col_idx]
                        if isinstance(quantite, (int, float)) and quantite > 0:
                            data.append({
                                'type_repas': current_type,
                                'categorie': categorie,
                                'jour_semaine': jour.title(),
                                'quantite': int(quantite)
                            })
                    except (IndexError, ValueError, TypeError):
                        continue
        
        return data
    
    def save_to_database(self, data: List[Dict]):
        """Sauvegarde les données dans la base"""
        if not self.db_connection:
            raise ValueError("Base de données non initialisée")
        
        cursor = self.db_connection.cursor()
        
        for record in data:
            cursor.execute('''
                INSERT INTO commandes (
                    ecole, ecole_normalized, agent, jour_semaine,
                    type_repas, categorie, quantite, source_fichier,
                    date_extraction
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                record.get('ecole'),
                record.get('ecole_normalized'),
                record.get('agent'),
                record.get('jour_semaine'),
                record.get('type_repas'),
                record.get('categorie'),
                record.get('quantite'),
                record.get('source_fichier'),
                datetime.now().isoformat()
            ))
        
        self.db_connection.commit()
    
    def close_database(self):
        """Ferme la connexion à la base"""
        if self.db_connection:
            self.db_connection.close()


class CrediExtractorGUI:
    """Interface graphique de l'extracteur CREDI"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("CREDI Extractor v1.0 - Extraction de commandes de repas")
        self.root.geometry("900x700")
        
        # Variables
        self.folder_path = tk.StringVar()
        self.mapping_file = tk.StringVar()
        self.db_path = None
        
        # Normalizer et Extractor
        self.normalizer = SchoolNameNormalizer()
        self.extractor = None
        
        # Style
        self.setup_styles()
        
        # Interface
        self.create_widgets()
    
    def setup_styles(self):
        """Configure les styles ttk"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Couleurs
        style.configure('Title.TLabel', font=('Helvetica', 16, 'bold'), foreground='#1f2937')
        style.configure('Subtitle.TLabel', font=('Helvetica', 10), foreground='#6b7280')
        style.configure('Section.TLabel', font=('Helvetica', 11, 'bold'), foreground='#374151')
        
        style.configure('Primary.TButton', font=('Helvetica', 10), padding=10)
        style.configure('Success.TButton', font=('Helvetica', 10), padding=10, background='#10b981')
    
    def create_widgets(self):
        """Crée les widgets de l'interface"""
        
        # En-tête
        header_frame = ttk.Frame(self.root, padding="20")
        header_frame.pack(fill=tk.X)
        
        ttk.Label(
            header_frame,
            text="🍽️ Extracteur CREDI",
            style='Title.TLabel'
        ).pack()
        
        ttk.Label(
            header_frame,
            text="Extraction automatique de commandes de repas depuis Excel et PDF",
            style='Subtitle.TLabel'
        ).pack()
        
        # Section mapping
        mapping_frame = ttk.LabelFrame(self.root, text="📝 Fichier de mapping (optionnel)", padding="10")
        mapping_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Label(
            mapping_frame,
            text="Pour normaliser automatiquement les noms d'écoles"
        ).pack(anchor=tk.W)
        
        mapping_input_frame = ttk.Frame(mapping_frame)
        mapping_input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Entry(
            mapping_input_frame,
            textvariable=self.mapping_file,
            state='readonly'
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        ttk.Button(
            mapping_input_frame,
            text="Charger mapping JSON",
            command=self.load_mapping
        ).pack(side=tk.LEFT)
        
        ttk.Button(
            mapping_input_frame,
            text="Télécharger modèle",
            command=self.download_template
        ).pack(side=tk.LEFT, padx=5)
        
        # Section sélection du dossier
        folder_frame = ttk.LabelFrame(self.root, text="📁 Dossier source", padding="10")
        folder_frame.pack(fill=tk.X, padx=20, pady=10)
        
        folder_input_frame = ttk.Frame(folder_frame)
        folder_input_frame.pack(fill=tk.X)
        
        ttk.Entry(
            folder_input_frame,
            textvariable=self.folder_path,
            state='readonly'
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        ttk.Button(
            folder_input_frame,
            text="Sélectionner dossier",
            command=self.select_folder,
            style='Primary.TButton'
        ).pack(side=tk.LEFT)
        
        # Barre de progression
        progress_frame = ttk.Frame(self.root, padding="20")
        progress_frame.pack(fill=tk.X)
        
        self.progress_label = ttk.Label(progress_frame, text="En attente...")
        self.progress_label.pack(anchor=tk.W)
        
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            mode='determinate',
            length=400
        )
        self.progress_bar.pack(fill=tk.X, pady=5)
        
        self.progress_detail = ttk.Label(progress_frame, text="0 / 0 fichiers")
        self.progress_detail.pack(anchor=tk.W)
        
        # Zone de log
        log_frame = ttk.LabelFrame(self.root, text="📋 Log de traitement", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            height=15,
            bg='#1f2937',
            fg='#d1d5db',
            font=('Courier', 9),
            wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # Boutons d'action
        action_frame = ttk.Frame(self.root, padding="20")
        action_frame.pack(fill=tk.X)
        
        self.btn_process = ttk.Button(
            action_frame,
            text="▶️ Traiter les fichiers",
            command=self.process_files,
            state=tk.DISABLED,
            style='Primary.TButton'
        )
        self.btn_process.pack(side=tk.LEFT, padx=5)
        
        self.btn_download = ttk.Button(
            action_frame,
            text="💾 Télécharger la base .db",
            command=self.download_database,
            state=tk.DISABLED,
            style='Success.TButton'
        )
        self.btn_download.pack(side=tk.LEFT, padx=5)
        
        self.btn_export_schools = ttk.Button(
            action_frame,
            text="📤 Exporter liste des écoles",
            command=self.export_schools,
            state=tk.DISABLED
        )
        self.btn_export_schools.pack(side=tk.LEFT, padx=5)
        
        self.btn_mapping_editor = ttk.Button(
            action_frame,
            text="✏️ Éditeur de mapping",
            command=self.open_mapping_editor
        )
        self.btn_mapping_editor.pack(side=tk.LEFT, padx=5)
    
    def log(self, message: str, level: str = 'INFO'):
        """Ajoute un message au log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        prefix = {
            'INFO': '✓',
            'ERROR': '✗',
            'WARNING': '⚠',
            'SUCCESS': '✓'
        }.get(level, 'ℹ')
        
        self.log_text.insert(tk.END, f"[{timestamp}] {prefix} {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
    
    def load_mapping(self):
        """Charge un fichier de mapping"""
        filename = filedialog.askopenfilename(
            title="Sélectionner le fichier de mapping",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if filename:
            if self.normalizer.load_mapping_file(filename):
                self.mapping_file.set(filename)
                self.log(f"Mapping chargé: {os.path.basename(filename)}", 'SUCCESS')
                messagebox.showinfo("Succès", "Fichier de mapping chargé avec succès!")
            else:
                self.log(f"Erreur lors du chargement du mapping", 'ERROR')
                messagebox.showerror("Erreur", "Impossible de charger le fichier de mapping")
    
    def download_template(self):
        """Télécharge un modèle de fichier de mapping"""
        template = {
            "version": "1.0",
            "description": "Template de mapping pour normalisation des noms d'écoles",
            "replacements": [
                {"pattern": "\\bSTE\\b", "replace": "SAINTE"},
                {"pattern": "\\bST\\b", "replace": "SAINT"},
                {"pattern": "\\bMAT\\b", "replace": "MATERNELLE"},
                {"pattern": "\\bELEM\\b", "replace": "ELEMENTAIRE"}
            ],
            "mappings": {
                "_exemple_1": "STE AURELIE MAT -> SAINTE AURELIE",
                "STE AURELIE MAT": "SAINTE AURELIE",
                "SAINTE AURELIE MATERNELLE": "SAINTE AURELIE"
            }
        }
        
        filename = filedialog.asksaveasfilename(
            title="Enregistrer le modèle",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json")],
            initialfile="credi_school_mappings_template.json"
        )
        
        if filename:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(template, f, ensure_ascii=False, indent=2)
            
            self.log(f"Modèle téléchargé: {os.path.basename(filename)}", 'SUCCESS')
            messagebox.showinfo("Succès", f"Modèle enregistré:\n{filename}")
    
    def select_folder(self):
        """Sélectionne le dossier source"""
        folder = filedialog.askdirectory(title="Sélectionner le dossier contenant les fichiers")
        
        if folder:
            self.folder_path.set(folder)
            self.btn_process.config(state=tk.NORMAL)
            self.log(f"Dossier sélectionné: {folder}", 'SUCCESS')
    
    def process_files(self):
        """Traite les fichiers du dossier"""
        if not self.folder_path.get():
            messagebox.showerror("Erreur", "Veuillez sélectionner un dossier")
            return
        
        # Désactiver les boutons
        self.btn_process.config(state=tk.DISABLED)
        
        # Lancer le traitement dans un thread
        thread = threading.Thread(target=self._process_files_thread)
        thread.start()
    
    def _process_files_thread(self):
        """Thread de traitement des fichiers"""
        try:
            folder = self.folder_path.get()
            
            # Créer la base de données
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.db_path = os.path.join(folder, f"credi_commandes_{timestamp}.db")
            
            self.extractor = MealOrderExtractor(self.normalizer)
            self.extractor.create_database(self.db_path)
            
            self.log("Base de données créée", 'SUCCESS')
            
            # Lister les fichiers
            files = []
            for ext in ['*.xlsx', '*.xls', '*.pdf']:
                files.extend(Path(folder).rglob(ext))
            
            total_files = len(files)
            self.log(f"Fichiers trouvés: {total_files}", 'INFO')
            
            if total_files == 0:
                self.log("Aucun fichier Excel ou PDF trouvé", 'WARNING')
                return
            
            # Traiter chaque fichier
            processed = 0
            errors = 0
            total_records = 0
            
            for i, file_path in enumerate(files, 1):
                try:
                    self.progress_label.config(text=f"Traitement: {file_path.name}")
                    self.progress_bar['value'] = (i / total_files) * 100
                    self.progress_detail.config(text=f"{i} / {total_files} fichiers")
                    
                    self.log(f"Traitement de: {file_path.name}", 'INFO')
                    
                    # Extraction selon le type de fichier
                    if file_path.suffix.lower() in ['.xlsx', '.xls']:
                        data = self.extractor.extract_from_excel(str(file_path))
                    elif file_path.suffix.lower() == '.pdf':
                        data = self.extractor.extract_from_pdf(str(file_path))
                    else:
                        continue
                    
                    if data:
                        self.extractor.save_to_database(data)
                        self.log(f"  ✓ {len(data)} enregistrements extraits", 'SUCCESS')
                        total_records += len(data)
                        processed += 1
                    else:
                        self.log(f"  ⚠ Aucune donnée extraite", 'WARNING')
                
                except Exception as e:
                    self.log(f"  ✗ Erreur: {str(e)}", 'ERROR')
                    errors += 1
            
            # Finaliser
            self.extractor.close_database()
            
            self.log("=" * 60, 'INFO')
            self.log(f"Traitement terminé!", 'SUCCESS')
            self.log(f"Fichiers traités: {processed}/{total_files}", 'INFO')
            self.log(f"Erreurs: {errors}", 'INFO' if errors == 0 else 'WARNING')
            self.log(f"Total enregistrements: {total_records}", 'SUCCESS')
            self.log(f"Base de données: {os.path.basename(self.db_path)}", 'SUCCESS')
            
            # Rapport de normalisation
            report = self.normalizer.get_normalization_report()
            if report:
                self.log("=" * 60, 'INFO')
                self.log("RAPPORT DE NORMALISATION:", 'INFO')
                self.log(report, 'INFO')
            
            # Activer les boutons de téléchargement
            self.btn_download.config(state=tk.NORMAL)
            self.btn_export_schools.config(state=tk.NORMAL)
            
            self.progress_label.config(text="✓ Traitement terminé!")
            
            messagebox.showinfo(
                "Succès",
                f"Traitement terminé!\n\n"
                f"Fichiers traités: {processed}/{total_files}\n"
                f"Enregistrements: {total_records}\n"
                f"Base créée: {os.path.basename(self.db_path)}"
            )
        
        except Exception as e:
            self.log(f"Erreur critique: {str(e)}", 'ERROR')
            messagebox.showerror("Erreur", f"Erreur lors du traitement:\n{str(e)}")
        
        finally:
            self.btn_process.config(state=tk.NORMAL)
    
    def download_database(self):
        """Télécharge la base de données"""
        if not self.db_path or not os.path.exists(self.db_path):
            messagebox.showerror("Erreur", "Aucune base de données disponible")
            return
        
        dest = filedialog.asksaveasfilename(
            title="Enregistrer la base de données",
            defaultextension=".db",
            filetypes=[("SQLite Database", "*.db")],
            initialfile=os.path.basename(self.db_path)
        )
        
        if dest:
            import shutil
            shutil.copy2(self.db_path, dest)
            self.log(f"Base téléchargée: {dest}", 'SUCCESS')
            messagebox.showinfo("Succès", f"Base de données enregistrée:\n{dest}")
    
    def export_schools(self):
        """Exporte la liste des écoles trouvées"""
        if not self.normalizer.found_schools:
            messagebox.showwarning("Attention", "Aucune école à exporter. Traitez d'abord les fichiers.")
            return
        
        folder = filedialog.askdirectory(title="Sélectionner le dossier de destination")
        
        if folder:
            json_file, txt_file = self.normalizer.export_found_schools(folder)
            self.log(f"Écoles exportées:", 'SUCCESS')
            self.log(f"  - {os.path.basename(json_file)}", 'INFO')
            self.log(f"  - {os.path.basename(txt_file)}", 'INFO')
            
            messagebox.showinfo(
                "Succès",
                f"Liste des écoles exportée:\n\n"
                f"JSON: {os.path.basename(json_file)}\n"
                f"TXT: {os.path.basename(txt_file)}"
            )
    
    def open_mapping_editor(self):
        """Ouvre l'éditeur de mapping"""
        MappingEditorWindow(self.root, self.normalizer)


class MappingEditorWindow:
    """Fenêtre d'édition du mapping"""
    
    def __init__(self, parent, normalizer: SchoolNameNormalizer):
        self.window = tk.Toplevel(parent)
        self.window.title("Éditeur de Mapping")
        self.window.geometry("800x600")
        
        self.normalizer = normalizer
        self.session_mappings = {}
        
        self.create_widgets()
    
    def create_widgets(self):
        """Crée les widgets de l'éditeur"""
        
        # Instructions
        info_frame = ttk.Frame(self.window, padding="10")
        info_frame.pack(fill=tk.X)
        
        ttk.Label(
            info_frame,
            text="💡 Créez des correspondances pour normaliser les noms d'écoles",
            font=('Helvetica', 10, 'bold')
        ).pack(anchor=tk.W)
        
        # Ajout de mapping
        add_frame = ttk.LabelFrame(self.window, text="Ajouter une correspondance", padding="10")
        add_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(add_frame, text="Nom original:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.original_entry = ttk.Entry(add_frame, width=40)
        self.original_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(add_frame, text="Nom normalisé:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.canonical_entry = ttk.Entry(add_frame, width=40)
        self.canonical_entry.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Button(
            add_frame,
            text="➕ Ajouter",
            command=self.add_mapping
        ).grid(row=0, column=2, rowspan=2, padx=5)
        
        # Liste des mappings
        list_frame = ttk.LabelFrame(self.window, text="Mappings de session", padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.mapping_listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            font=('Courier', 9)
        )
        self.mapping_listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.mapping_listbox.yview)
        
        # Boutons d'action
        action_frame = ttk.Frame(self.window, padding="10")
        action_frame.pack(fill=tk.X)
        
        ttk.Button(
            action_frame,
            text="💾 Télécharger mapping JSON",
            command=self.download_mapping
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            action_frame,
            text="🗑️ Effacer tout",
            command=self.clear_mappings
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            action_frame,
            text="Fermer",
            command=self.window.destroy
        ).pack(side=tk.RIGHT, padx=5)
    
    def add_mapping(self):
        """Ajoute un mapping"""
        original = self.original_entry.get().strip().upper()
        canonical = self.canonical_entry.get().strip().upper()
        
        if not original or not canonical:
            messagebox.showerror("Erreur", "Veuillez remplir les deux champs")
            return
        
        # Supprimer les accents de l'original
        import unicodedata
        original_normalized = ''.join(
            c for c in unicodedata.normalize('NFD', original)
            if unicodedata.category(c) != 'Mn'
        )
        
        self.session_mappings[original_normalized] = canonical
        
        self.mapping_listbox.insert(tk.END, f"{original} → {canonical}")
        
        self.original_entry.delete(0, tk.END)
        self.canonical_entry.delete(0, tk.END)
    
    def download_mapping(self):
        """Télécharge le mapping au format JSON"""
        if not self.session_mappings:
            messagebox.showwarning("Attention", "Aucun mapping à télécharger")
            return
        
        mapping_data = {
            "version": "custom",
            "description": "Mapping généré par l'éditeur CREDI",
            "date_generated": datetime.now().isoformat(),
            "replacements": self.normalizer.replacements,
            "mappings": dict(self.normalizer.mappings, **self.session_mappings)
        }
        
        filename = filedialog.asksaveasfilename(
            title="Enregistrer le mapping",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json")],
            initialfile=f"credi_school_mappings_{datetime.now().strftime('%Y%m%d')}.json"
        )
        
        if filename:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(mapping_data, f, ensure_ascii=False, indent=2)
            
            messagebox.showinfo("Succès", f"Mapping enregistré:\n{filename}")
    
    def clear_mappings(self):
        """Efface tous les mappings de session"""
        if messagebox.askyesno("Confirmation", "Effacer tous les mappings de session?"):
            self.session_mappings.clear()
            self.mapping_listbox.delete(0, tk.END)


def main():
    """Point d'entrée principal"""
    
    # Vérifier les dépendances
    missing_deps = []
    
    if openpyxl is None:
        missing_deps.append("openpyxl")
    
    if pdfplumber is None:
        missing_deps.append("pdfplumber")
    
    if missing_deps:
        root = tk.Tk()
        root.withdraw()
        messagebox.showwarning(
            "Dépendances manquantes",
            f"Les bibliothèques suivantes sont manquantes:\n\n" +
            "\n".join(f"- {dep}" for dep in missing_deps) +
            f"\n\nInstallez-les avec:\npip install {' '.join(missing_deps)}"
        )
        root.destroy()
    
    # Lancer l'application
    root = tk.Tk()
    app = CrediExtractorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
